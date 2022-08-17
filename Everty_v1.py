import openpyxl # This is a library that functions with excel files
from math import sqrt # A simple Math library
from openpyxl.utils import get_column_letter # calling specifically the column number exchange function
import itertools # This is a library that helps iteration
import pickle # This is a library that helps the creation of file data objects
import sys # This is a system library 
import os # This is a system library 
from datetime import datetime as dt # This is a date and time library 

__author__ = "Nicholas Rademaker"
__copyright__ = None
__credits__ = ["Nicholas Rademaker"]
__license__ = None
__version__ = "2.0.0"
__maintainer__ = "Nicholas Rademaker"
__email__ = "nicholas.rademaker@mastercard.com"
__status__ = "Development"

class N():
    def __init__(self, obj, **kwarg): #kwarg is a contraction of key-word-arguement
        '''
        This main class forms the basis of other data information files
        '''
        self.obj = obj
        if 'name' in kwarg.keys():
            self.name = kwarg['name']
        else:
            self.name = ''
        if 'nr' in kwarg.keys():
            self.kw = kwarg['nr']
        else:
            self.kw = 1
            
            
    def c_report(self, blob, cn1, cn2):
        '''
        Creates a report on the correlation of variables in an array 
        Blob is the dictionary object    
        '''
        if type(cn1) != list:
            cn1 = [i[cn1] for i in blob]
            cn2 = [i[cn2] for i in blob]
        list_1 = self._norm(cn1)
        list_2 = self._norm(cn2)
        return pearsonr(list_1, list_2)

    def general_search(self, blob, params=False):
        d = {}
        out = {}
        if params:
            for p in params:
                d[p] = [i[p] for i in blob]
            for s in itertools.combinations(params, 2):
                out['-'.join(s)] = c_report(blob, d[s[0]], d[s[1]])

    def strain(blob):
        f = list(filter(lambda x: x["Gender"] == 'female',
                        blob))
        fp = list(filter(lambda x: x["Gender"] == 'female' and x["Location Name"] == "Pune India",
                         blob))
        mp = list(filter(lambda x: x["Gender"] == 'male' and x["Location Name"] == "Pune India",
                         blob))
        return f, fp, mp

    def ave(blob):
        b = {}
        for pers in blob:
            for key, value in pers.items():
                if value:
                    if type(value) == int:
                        if not key in b.keys():
                            b[key] = [0, 0]
                        b[key][0] += int(value)
                        b[key][1] += 1
        return {key: value[0] / value[1] for key, value in b.items()}

    def mode(blob):
        b = {}
        for pers in blob:
            for key, value in pers.items():
                if value:
                    if type(value) == int:
                        if not key in b.keys():
                            b[key] = []
                        b[key] += [value]
        for key in b.keys():
            b[key] = max(set(b[key]), key=b[key].count)
        return b

    def std(blob):
        b = {}
        for pers in blob:
            for key, value in pers.items():
                if value:
                    if type(value) == int:
                        if not key in b.keys():
                            b[key] = []
                        b[key] += [value]
        for key in b.keys():
            lst = b[key]
            num_items = len(lst)
            mean = sum(lst) / num_items
            differences = [x - mean for x in lst]
            sq_differences = [d ** 2 for d in differences]
            ssd = sum(sq_differences)
            variance = ssd / num_items
            b[key] = sqrt(variance)
        return b

    def panalyze(blob):
        def analyze(blob):
            mi = mode(blob)
            a = ave(blob)
            ma = std(blob)
            return mi,a,ma
        print("%s~%s~%s" % (mi, a, ma))

    def _norm(self, array):
        return [float(i) / sum(array) for i in array]

    def _current_line(self):
        '''
        Gives the current bottom line of the excel DB
        '''
        if self.kw:
            line = int(self.kw)
        else:
            line = 1
        while self.sheet['A%s' % (line)].value is not None:
            line += 1

        return line
        
class Ncreate(N):
    def __init__(self, obj, **kwarg): 
        '''        
        This class should handel the creation of excel files from a pickle created data object
        '''
        N.__init__(self, obj, **kwarg)
        self.wb = openpyxl.Workbook()
        self.filepath = 'defaults' # should replace with input for file name
        if 'name' in kwarg.keys():
            self.filepath = kwarg['name']
      
       
    def create(self):
        with open(self.obj ,'rb') as file:
            try:
                obj = pickle.load(file)
            except Exception as ex:
                print('Exception: "%s" has occured exiting'%ex)
                exit()

        '''
        for key in data.keys():
            for k, v in data[key].items():
                if key == v:
                    self.ucolumn = key
                    break
            break
        '''

        
        # Greater iterator:

        gobj = obj
        for sheet, obj in gobj.items():
            self.sheet = self.wb.create_sheet(sheet)
            names = list(obj[0].keys())
            for name in names:
                try:
                    if self.kw:
                        self.sheet['%s%s' % (get_column_letter(names.index(name)+1), self.kw)] = name
                except KeyError:
                    print("Key error has occured")
                    print("%s  %s" % (get_column_letter(names.index(name)), name))
                    exit()
            print('uploading..')
            tots = len(obj)
            c = 0
            for col in names:
                self.sheet['%s%s' %
                                       (get_column_letter(names.index(col)+1), self.kw)] = col
         
            for key in range(1, tots):
                for col in names:
                    c += 1
                    print('Currently at %s out of %s or at %s%%' %
                          (key, tots, (key / tots) * 100), end='\r')
                    try:
                        if not type(obj[key][col]) == None:
                            self.sheet['%s%s' %
                                       (get_column_letter(names.index(col)+1), key+self.kw)] = obj[key-self.kw][col]
                    except KeyError:
                        print("a Key error has occurred in _commit")
                        print(key)
                        print(col)
                    except TypeError as ex:
                        print('TypeError: %s' % ex)
                        exit()
        #final line
        self.commit_to_xlxs()
    
    def _sheet_load(self):
        '''
        Loads sheets into dictionary db
        '''
        ncolumns = self._column_scan()
        self.db[self.sheetn]["Meta"]["column names"] = ncolumns
        cd = {get_column_letter(x + 1): ncolumns[x]
              for x in range(len(ncolumns))}
        cdr = dict(map(reversed, cd.items()))
        scanlist = self.scan_list(self.unique)
        uidd = {x[0]: x[1] for x in scanlist}
        self.db[self.sheetn]["Meta"]["column dic"] = cd
        self.db[self.sheetn]["Meta"]["column dic r"] = cdr
        self.db[self.sheetn]["Meta"]["uid key dic"] = uidd    
        
    def update(self, dick):
        '''
        Must be at two story dictionary object with a unique column
        Adds to current dictionary object
        '''
        assert isinstance(dick, dict), "Object is not a dictionary"
        names = self.db[self.sheetn]["Meta"]["column names"]
        for key, value in dick.items():  # (key : value={subkey:subvalue})
            for x in value.keys():  # x=subkey
                if x not in names:  # if subkey not in column name
                    self.add_column(x)  # if not add it
            # insert it into the data base by applying it's key to its value
            self.db[self.sheetn]['Data'][key] = value
            self._update_uid(key)  # then I add the unique ID key
        #  print(self.db[self.sheetn]['Data']) thingy
        k = all([key in self.db[self.sheetn]['Data'].keys()
                 for key in dick.keys()])
        self.lastdick = dick
        if not k:
            self.update(dick)
            
            
    
    def commit_to_xlxs(self):
        #self._commit()  # EDITED WARNING WARNING WARNING
        #with open(self.filepath+'key.pickle', 'wb') as file:
        #    pickle.dump(self.hashtable, file)
        self.wb.save(self.filepath + '.xlsx')

    def _commit(self):
        dick = self.db[self.sheetn]["Meta"]["column dic r"]
        names = self.db[self.sheetn]["Meta"]["column names"]
        data = self.db[self.sheetn]['Data']
        uidic = self.db[self.sheetn]["Meta"]["uid key dic"]
        conv = self.db[self.sheetn]["Meta"]["column dic r"]
        for key in data.keys():
            for k, v in data[key].items():
                if key == v:
                    self.ucolumn = key
                    break
            break

        for name in names:
            try:
                self.sheet['%s%s' % (conv[name], 1)] = name
            except KeyError:
                print("Key error has occured")
                print("%s  %s" % (conv, name))
                exit()
        print('uploading..')
        tots = len(data.keys()) * len(names)
        c = 0
        for key in data.keys():
            for col in names:
                c += 1
                print('Currently at %s out of %s or at %s%%' %
                      (c, tots, (c / tots) * 100), end='\r')
                try:
                    if data[key][col]:
                        if isinstance(data[key][col], bytes):
                            ba = bytearray(data[key][col])
                            b = '-'.join([str(x) for x in list(ba)])
                        else:
                            b = str(data[key][col])
                        self.sheet['%s%s' %
                                   (dick[col], uidic[key])] = b
                except KeyError:
                    print("a Key error has occurred in _commit")
                    print(key)
                    print(col)
                    self.update(self.lastdick)
                except TypeError as ex:
                    print('TypeError: %s' % ex)
                    print(data[key][col])
                    exit()

        self._sheet_load()
        
     
    def _current_width(self):
        '''
        Gives the current side line of the excel DB
        '''
        line = 1
        while self.sheet['%s%s' % (get_column_letter(line), self.idc)].value is not None:
            line += 1
        return line

    def _column_scan(self):
        '''
        Gives the full list of column names
        '''
        col = 1
        indexes = []
        while self.sheet['%s%s' % (get_column_letter(col), self.idc)].value is not None:
            indexes += [self.sheet['%s%s' % (get_column_letter(col), self.idc)].value]
            col += 1
        return indexes

        # Utility

    def scan_list(self, ind):  # Make discrete and change name
        you = []
        cl = self.db[self.sheetn]["Meta"]["bottom"]
        s = 1
        for x in range(s, cl):
            if self.sheet['%s%s' % (ind, x)].value:
                inser = (self.sheet['%s%s' % (ind, x)].value, x)
            else:
                inser = (str(x - 1), x)
            you.append(inser)
        return you

    def _buf(self, string, space):
        buf = int((space - len(string)) / 4)
        sbuf = ' ' * buf
        return sbuf + string + sbuf


    def self_scan(self):
        db = {x: dict() for x in
              self.db[self.sheetn]["Meta"]["uid key dic"].keys()}
        for x in db.keys():
            for y in self.db[self.sheetn]["Meta"]["column names"]:
                try:
                    db[x][y] =\
                        self.sheet['%s%s' %
                                   (self.db[self.sheetn]["Meta"]["column dic r"][y],
                                    self.db[self.sheetn]["Meta"]["uid key dic"][x])].value
                except:
                    print("an error occured")
                    print(self.db[self.sheetn]["Meta"]["column dic r"])
                    print(self.db[self.sheetn]["Meta"]["column names"])
                    print(y)
        return db

        
        
    def normalize(self, blob, c_names):
        '''
        For each individual in list
        for each relation of the individual
        check if the value of that relationship is "numerical"\
        if yes stop and ignore
        if no add to set and continue
        when finished check if non-normalized data has "nearness"
        if yes get user to label options
        if no randomly assign.
        '''
        meta = {c: {} for c in c_names}
        count = 1
        for c in c_names:
            if type(blob[1][c]) in (int, float) or blob[1][c].isdigit():
                continue
            yorn = True if input(
                'Does data: %s have "nearness" (y/n)?: ' % c) == 'y' else False
            for i in range(len(blob)):
                if yorn:
                    s = int(input('%s value?: ' % blob[i][c].__str__()))
                    if blob[i][c].__str__() in list(meta[c].values()):
                        continue
                    else:
                        meta[c][s] = blob[i][c].__str__()
                    blob[i][c] = s
                else:
                    meta[c][count] = blob[i][c].__str__()
                    blob[i][c] = count
                    count += 1
        with open(input('Name of object?: '), 'wb') as file:
            pickle.dump([meta, blob], file)

class N_weekly(N):
    def __init__(self, obj, obj2, **kwarg): 
        '''        
        self.obj = obj
        if kwarg['name']:
            self.name = kwarg['name']
        if kwarg['nr']:
            self.kw = kwarg['nr']
        '''
        N.__init__(self, obj, **kwarg)
        self.wb = openpyxl.Workbook()
        self.filepath = 'defaults'
        if 'name' in kwarg.keys():
            self.filepath = kwarg['name']  
            
        self.obj_1 = obj # This should be the old record file
        self.obj_2 = obj2 # this should be the update sheet
        
        self.tt = True
        

    def create(self):
        with open(self.obj_1 ,'rb') as file:
            try:
                old_record = pickle.load(file)
            except Exception as ex:
                print('Exception: "%s" has occured exiting'%ex)
                exit()
                
        with open(self.obj_2 ,'rb') as file:
            try:
                update = pickle.load(file)
            except Exception as ex:
                print('Exception: "%s" has occured exiting'%ex)
                exit()
        
        # Update reference sheet
        key = str(dt.now().day) +'-' + (str(dt.now().month) if len == 2 else '0' + str(dt.now().month))
        
        
        old_record[key] = old_record.pop(list(old_record.keys())[3])
        rw = [0,1,-1,2]
        new_record = {}
        for x in rw:
            new_record[list(old_record.keys())[x]] = old_record[list(old_record.keys())[x]]
        up = [(x['DPAN Start Range'],x['Total Token Count']) for x in update[list(update.keys())[0]]]
        
        new_record[key] = [{'DPAN Start Range':x[0],'Total Token Count':x[1]} for x in up]

        print('New update sheet recorded')
        print('Starting overview update')
        save = new_record[key] # Patch resolve later
        names = list(new_record['Overview'][0].keys())
        # new_names keeps the names of the columns
        new_names = names[:10] + names[11:13] + [dt.now()] + names[14:16] + ['Days Remaining '+ dt.now().strftime('%d/%m/%Y')] + names[16:]
        old = new_record['Overview']
        new_record = []
        forms = []
        c = 2
        '''
        Error to handel:
        Traceback (most recent call last):
        File "evert.py", line 614, in <module>
            app.create()
        File "evert.py", line 461, in create
            rec[new_names[n]] = up[r][1]
        IndexError: list index out of range
        '''
        
        
        
        for r in range(len(old)):
            rec = {}
            for n in range(len(new_names)):
                #rec[new_names[n]] = up[r][1] 
                if n in [8,9,12,15,16]:
                    if n == 8:
                       rec[new_names[n]] = '=MAX(0,H%s-INDIRECT(CONCAT(RIGHT($I$1,1),ROW())))'%str(r+2)
                       rec[new_names[n]] = rec[new_names[n]].replace('@','')
                    if n == 9:
                        rec[new_names[n]] = '=1-(I%s/H%s)'%(str(r+2),str(r+2))
                    if n == 12:
                        rec[new_names[n]] = up[r][1]
                    if n == 15:
                        rec[new_names[n]] = '=IF(M%s-L%s=0,0,IF($I%s/((M%s-L%s)/(M$1-L$1))<0,0,($I%s/((M%s-L%s)/(M$1-L$1)))))'%(r+2,r+2,r+2,r+2,r+2,r+2,r+2,r+2)
                        #IF(M2-L2=0,0,IF($I2/((M2-L2)/(M$1-L$1))<0,0,($I2/((M2-L2)/(M$1-L$1)))))
                        # This is a complex inline analogue of James Noe's original token calculation
                        #########################################################################
                        '''
                        dif = old[r][new_names[11]] - old[r][new_names[10]]
                        try:
                            res = old[r][new_names[8]]/(dif/(int(new_names[11].strftime('%d'))- int(new_names[10].strftime('%d'))))
                        except ZeroDivisionError as ex:
                            rec[new_names[n]] = 0
                        except TypeError as ex:
                            rec[new_names[n]] = 0

                            
                        rec[new_names[n]] = 0 if (dif == 0) or (res < 0) else res
                        '''
                        #elif old[r][new_names[8]]/(up[r][1] - old[r][new_names[11]]/int(str(new_names[11]-new_names[10])[0])) < 0:

                    if n == 16:
                        rec[new_names[n]] = '=IF(P%s-O%s>100,"Yes","No")'%(r+2,r+2)
                else:
                    rec[new_names[n]] = old[r][new_names[n]]
            new_record += [rec]
        #print(new_record) This should be the completed new records
        old_record['Overview'] = new_record
        old_record[key] = save
        with open('NEW!_Report','wb') as file:
            pickle.dump(old_record, file)
            
        print('Pickle file record completed')
                
        '''
        # Greater iterator:
        gobj = obj
        for sheet, obj in gobj.items():
            self.sheet = self.wb.create_sheet(sheet)
            names = list(obj[0].keys())
            for name in names:
                try:
                    if self.kw:
                        self.sheet['%s%s' % (get_column_letter(names.index(name)+1), self.kw)] = name
                except KeyError:
                    print("Key error has occured")
                    print("%s  %s" % (get_column_letter(names.index(name)), name))
                    exit()
            print('uploading..')
            tots = len(obj)
            c = 0
            for col in names:
                self.sheet['%s%s' %
                                       (get_column_letter(names.index(col)+1), self.kw)] = col
         
            for key in range(1, tots):
                for col in names:
                    c += 1
                    print('Currently at %s out of %s or at %s%%' %
                          (key, tots, (key / tots) * 100), end='\r')
                    try:
                        if obj[key][col]:
                            self.sheet['%s%s' %
                                       (get_column_letter(names.index(col)+1), key+self.kw)] = obj[key-self.kw][col]
                    except KeyError:
                        print("a Key error has occurred in _commit")
                        print(key)
                        print(col)
                    except TypeError as ex:
                        print('TypeError: %s' % ex)
                        exit()
        #final line
        self.commit_to_xlxs()
        '''

class Ndstruct(N):
    def __init__(self, obj, **kwarg):
        N.__init__(self, obj, **kwarg)
        print('Loading wb..')
        self.wb = openpyxl.load_workbook(self.obj,data_only=True,read_only=True) #added data_only to resolve formulas

        #self.wb = openpyxl.load_workbook(self.obj, data_only=True) #added data_only to resolve formulas

        self.sheet = self.wb.active
        
        if 'sheet' in kwarg.keys():
            self.shit = kwarg['sheet']
        else:
            self.shit = ''
            
            
    def batcher(self):
        '''collects pickles for jar'''
        picklejar = {}
        c = 0
        for sheet in self.wb.sheetnames:
            c += 1
            self.sheet = self.wb[sheet]
            picklejar[sheet] = self.tokenizer()
        
        return picklejar
            
            
    def dump(self, answered = False):
                        
        if self.name:
            pickle.dump(self.batcher(), open(self.name, 'wb'))
        elif self.shit:
            pickle.dump(self.batcher(), open(self.shit, 'wb'))
        else:
            inny = answered
            while not answered:
                inny = input('Name of object?: ')
                if inny:
                    answered = True
                else:
                    print('No answer given please try again')
            
            pickle.dump(self.batcher(), open(inny, 'wb'))
    def tokenizer(self):
        col = True
        x = 1
        relations = []
        while col:
            
            namerow = "%s" + str(self.kw)
            col = self.sheet[namerow % (get_column_letter(x))].value
            if col:
                relations += [col]
            x += 1
        print("Column list established")
        val = True
        num = self._current_line()
        wf = []
        for n in range(2, num + 1):
            wf += [{relations[y - 1]:self.sheet['%s%s' %
                                                (get_column_letter(y), n)].value for y in range(1, len(relations) + 1)}]
            print("%s out of %s" % (n, num), end='\r')
            
   
        return wf

if __name__ == "__main__":
    print('Arguements: %s'%sys.argv)
    if sys.argv[1] == 'Q':
        print('Weekly proper engage')
        app = Ndstruct('Account_Range_Capacity_V3.0.xlsx')
        app.dump("report")
        print('Report object created')
        app = Ndstruct(sorted(os.listdir())[2]) # change to 3
        app.dump("analysis")
        print('Analysis object created')
        app = N_weekly('analysis','report')
        print('Weekly report engaged')
        app.create()
        print('Weekly report completed')
        
    elif sys.argv[1][-5:] == '.xlsx' or sys.argv[1][-5:] == '.xlsb':
        app = Ndstruct(sys.argv[1]) # needs notes
        app.dump()
    elif len(sys.argv) == 3:
        print('Weekly initialized')
        app = N_weekly(sys.argv[1], sys.argv[2])
        app.create()
    else:
        if len(sys.argv) == 2:
            app = Ncreate(sys.argv[1])
        elif len(sys.argv) == 3:
            app = Ncreate(sys.argv[1], name=sys.argv[2])
        app.create()

'''
        if kwarg['name']:
            self.name = kwarg['name']
        if kwarg['nr']:
            self.kw = kwarg['nr']
'''
            
'''
if __name__ == "__main__":
    print('Arguements: %s'%sys.argv)
    if len(sys.argv) == 3:
        print('Weekly initialized')
        app = N_weekly(sys.argv[1], sys.argv[2])
        app.create()
    elif sys.argv[1][-5:] == '.xlsx'or sys.argv[1][-5:] == '.xlsb':
        app = Ndstruct(sys.argv[1])
        app.dump()
    else:
        if len(sys.argv) == 2:
            app = Ncreate(sys.argv[1])
        elif len(sys.argv) == 3:
            app = Ncreate(sys.argv[1], name=sys.argv[2])

        app.create()
        '''
