import openpyxl
from math import sqrt
from openpyxl.utils import get_column_letter
from scipy.stats.stats import pearsonr
import itertools
import pickle
import sys
import os

'''
sometimes inappropriately places None Urgent issue

'''


class Ndstruct():
    def __init__(self):  # This takes so long maybe threading?
        pass

    def tokenizer(self, excel, name=False, sheet=None):
        #assert excel[-5:] == '.xlsx', 'Needs to be xlsx file'
        wb = openpyxl.load_workbook(excel)
        print('finished loading wb')
        if sheet:
            self.sheet = wb[sheet]
        else:
            self.sheet = wb.active
        # make list
        col = True
        x = 1
        relations = []
        while col:
            col = self.sheet["%s1" % (get_column_letter(x))].value
            if col:
                relations += [col]
            x += 1
        print("Column list established")
        val = True
        num = 0
        wf = []
        for n in range(2, self._current_line() + 1):
            wf += [{relations[y - 1]:self.sheet['%s%s' %
                                                (get_column_letter(y), n)].value for y in range(1, len(relations) + 1)}]
            print("%s out of %s" % (n, num), end='\r')
        if sheet:
            pickle.dump(wf, open(sheet, 'wb'))
        elif isinstance(name, str):
            pickle.dump(wf, open(name, 'wb'))
        else:
            pickle.dump(wf, open(input('Name of object?: '), 'wb'))

    def _current_line(self):
        '''
        Gives the current bottom line of the excel DB
        '''
        line = 1
        while self.sheet['A%s' % (line)].value is not None:
            line += 1

        return line

    def normalize(self, blob, c_names):
        '''
        For each individual in list
        for each relation of the individual
        check if the value of that relationship is "numerical"\
        if yes stop and ignore
        if no add to set and continue
        when finished check if non-normalized data has "nearness"
        if yes get user to label options
        if no randomly assign.
        '''
        meta = {c: {} for c in c_names}
        count = 1
        for c in c_names:
            if type(blob[1][c]) in (int, float) or blob[1][c].isdigit():
                continue
            yorn = True if input(
                'Does data: %s have "nearness" (y/n)?: ' % c) == 'y' else False
            for i in range(len(blob)):
                if yorn:
                    s = int(input('%s value?: ' % blob[i][c].__str__()))
                    if blob[i][c].__str__() in list(meta[c].values()):
                        continue
                    else:
                        meta[c][s] = blob[i][c].__str__()
                    blob[i][c] = s
                else:
                    meta[c][count] = blob[i][c].__str__()
                    blob[i][c] = count
                    count += 1
        with open(input('Name of object?: '), 'wb') as file:
            pickle.dump([meta, blob], file)

    def c_report(self, blob, cn1, cn2):
        '''
        Creates a report on the correlation of variables in an array 
        Blob is the dictionary object    
        '''
        if type(cn1) != list:
            cn1 = [i[cn1] for i in blob]
            cn2 = [i[cn2] for i in blob]
        list_1 = self._norm(cn1)
        list_2 = self._norm(cn2)
        return pearsonr(list_1, list_2)

    def general_search(self, blob, params=False):
        d = {}
        out = {}
        if params:
            for p in params:
                d[p] = [i[p] for i in blob]
            for s in itertools.combinations(params, 2):
                out['-'.join(s)] = c_report(blob, d[s[0]], d[s[1]])

    def strain(blob):
        f = list(filter(lambda x: x["Gender"] == 'female',
                        blob))
        fp = list(filter(lambda x: x["Gender"] == 'female' and x["Location Name"] == "Pune India",
                         blob))
        mp = list(filter(lambda x: x["Gender"] == 'male' and x["Location Name"] == "Pune India",
                         blob))
        return f, fp, mp

    def ave(blob):
        b = {}
        for pers in blob:
            for key, value in pers.items():
                if value:
                    if type(value) == int:
                        if not key in b.keys():
                            b[key] = [0, 0]
                        b[key][0] += int(value)
                        b[key][1] += 1
        return {key: value[0] / value[1] for key, value in b.items()}

    def mode(blob):
        b = {}
        for pers in blob:
            for key, value in pers.items():
                if value:
                    if type(value) == int:
                        if not key in b.keys():
                            b[key] = []
                        b[key] += [value]
        for key in b.keys():
            b[key] = max(set(b[key]), key=b[key].count)
        return b

    def std(blob):
        b = {}
        for pers in blob:
            for key, value in pers.items():
                if value:
                    if type(value) == int:
                        if not key in b.keys():
                            b[key] = []
                        b[key] += [value]
        for key in b.keys():
            lst = b[key]
            num_items = len(lst)
            mean = sum(lst) / num_items
            differences = [x - mean for x in lst]
            sq_differences = [d ** 2 for d in differences]
            ssd = sum(sq_differences)
            variance = ssd / num_items
            b[key] = sqrt(variance)
        return b

    def analyze(blob):
        mi = mode(blob)
        a = ave(blob)
        ma = std(blob)
        for key in a.keys():
            print("%s~%s~%s~%s" % (key, mi[key], a[key], ma[key]))

    def _norm(self, array):
        return [float(i) / sum(array) for i in array]


if __name__ == "__main__":
    if sys.argv[1][-5:] == '.xlsx':
        app = Ndstruct()
        app.tokenizer(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
